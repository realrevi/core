"""
============================================================
CORE - Cut Optimization & Reporting Engine v4.0
Python Backend with PyWebview - IMPROVED VERSION
============================================================
"""

import sys
import os
import json
import sqlite3
import re
import subprocess
import platform
from datetime import datetime
from typing import Optional, Dict, List, Any
from pathlib import Path

# ============================================================
# FROZEN/EXE PATH FIX
# ============================================================

def get_base_path() -> Path:
    """Get base path for resources (works for both dev and frozen exe)"""
    if getattr(sys, 'frozen', False):
        # Running as compiled exe
        return Path(sys._MEIPASS)
    return Path(__file__).parent

def get_app_dir() -> Path:
    """Get application directory (for data storage)"""
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent

# ============================================================
# IMPORTS WITH ERROR HANDLING
# ============================================================

# bcrypt for password hashing
try:
    import bcrypt
    BCRYPT_AVAILABLE = True
except ImportError:
    import hashlib
    BCRYPT_AVAILABLE = False
    print("Warning: bcrypt not available, using SHA256 fallback")

# Excel processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not available")

# PDF export - Devre dışı bırakıldı (kullanılmıyor)
# try:
#     from reportlab.lib import colors
#     from reportlab.lib.pagesizes import A4
#     from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
#     from reportlab.lib.styles import getSampleStyleSheet
#     REPORTLAB_AVAILABLE = True
# except ImportError:
#     REPORTLAB_AVAILABLE = False
REPORTLAB_AVAILABLE = False

# PyWebview
try:
    import webview
    WEBVIEW_AVAILABLE = True
except ImportError:
    WEBVIEW_AVAILABLE = False
    print("Error: pywebview not available. Install with: pip install pywebview")
    sys.exit(1)

# ============================================================
# CONFIGURATION
# ============================================================

class Config:
    """Application configuration"""
    APP_NAME = "CORE"
    APP_VERSION = "4.0"
    APP_TITLE = f"{APP_NAME} - Cut Optimization & Reporting Engine v{APP_VERSION}"

    # Window settings
    WINDOW_WIDTH = 1400
    WINDOW_HEIGHT = 900
    WINDOW_MIN_WIDTH = 1000
    WINDOW_MIN_HEIGHT = 700

    @staticmethod
    def get_base_path() -> Path:
        """Get base path for bundled resources"""
        return get_base_path()

    @staticmethod
    def get_app_dir() -> Path:
        """Get application directory"""
        return get_app_dir()

    @staticmethod
    def get_data_dir() -> Path:
        """Get data directory - AppData altında (yazma izni için)"""
        # Windows'ta AppData/Local altına yaz
        if sys.platform == 'win32':
            app_data = os.environ.get('LOCALAPPDATA', os.path.expanduser('~'))
            data_dir = Path(app_data) / "CORE" / "data"
        else:
            # Linux/Mac için home dizini
            data_dir = Path.home() / ".core" / "data"
        
        try:
            data_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            # Fallback: geçici dizin
            import tempfile
            data_dir = Path(tempfile.gettempdir()) / "CORE" / "data"
            data_dir.mkdir(parents=True, exist_ok=True)
        
        return data_dir

    @staticmethod
    def get_db_path() -> Path:
        """Get SQLite database path"""
        return Config.get_data_dir() / "core.db"

    @staticmethod
    def get_json_path(name: str) -> Path:
        """Get JSON file path"""
        return Config.get_data_dir() / f"core_{name}.json"

    @staticmethod
    def get_html_path() -> Path:
        """Get index.html path"""
        return Config.get_base_path() / "index.html"

# ============================================================
# DATABASE MANAGER
# ============================================================

class DatabaseManager:
    """SQLite database manager"""

    def __init__(self):
        self.db_path = str(Config.get_db_path())
        self._init_database()

    def _init_database(self):
        """Initialize database tables"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_no TEXT,
                    date TEXT,
                    file_name TEXT,
                    file_path TEXT,
                    output_path TEXT,
                    total_parts INTEGER,
                    material_count INTEGER,
                    type_count INTEGER,
                    body_data TEXT,
                    thin_data TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS stats (
                    id INTEGER PRIMARY KEY,
                    total_jobs INTEGER DEFAULT 0,
                    total_parts INTEGER DEFAULT 0,
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("INSERT OR IGNORE INTO stats (id) VALUES (1)")
            conn.commit()

    def add_history(self, job: Dict) -> bool:
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO history
                    (job_no, date, file_name, file_path, output_path,
                     total_parts, material_count, type_count, body_data, thin_data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    job.get('job_no', ''),
                    job.get('date', ''),
                    job.get('file_name', ''),
                    job.get('file_path', ''),
                    job.get('output_path', ''),
                    job.get('stats', {}).get('parts', 0),
                    job.get('stats', {}).get('materials', 0),
                    job.get('stats', {}).get('types', 0),
                    json.dumps(job.get('results', {}).get('body', [])),
                    json.dumps(job.get('results', {}).get('thin', []))
                ))

                cursor.execute("""
                    UPDATE stats SET
                        total_jobs = total_jobs + 1,
                        total_parts = total_parts + ?,
                        last_updated = CURRENT_TIMESTAMP
                    WHERE id = 1
                """, (job.get('stats', {}).get('parts', 0),))
                conn.commit()
                return True
        except Exception as e:
            print(f"Database error: {e}")
            return False

    def delete_history(self, job_ids: List[int]) -> Dict:
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                placeholders = ','.join('?' * len(job_ids))

                cursor.execute(f"""
                    SELECT SUM(total_parts) as total FROM history
                    WHERE id IN ({placeholders})
                """, job_ids)
                result = cursor.fetchone()
                parts_to_remove = result[0] if result[0] else 0

                cursor.execute(f"DELETE FROM history WHERE id IN ({placeholders})", job_ids)
                deleted_count = cursor.rowcount

                cursor.execute("""
                    UPDATE stats SET
                        total_jobs = total_jobs - ?,
                        total_parts = total_parts - ?,
                        last_updated = CURRENT_TIMESTAMP
                    WHERE id = 1
                """, (deleted_count, parts_to_remove))
                conn.commit()
                return {'success': True, 'deleted': deleted_count}
        except Exception as e:
            print(f"Database error: {e}")
            return {'success': False, 'error': str(e)}

    def get_history(self, limit: int = 100) -> List[Dict]:
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT * FROM history ORDER BY created_at DESC LIMIT ?
                """, (limit,))
                rows = cursor.fetchall()

                return [{
                    'id': row['id'],
                    'job_no': row['job_no'],
                    'date': row['date'],
                    'file_name': row['file_name'],
                    'file_path': row['file_path'],
                    'output_path': row['output_path'],
                    'stats': {
                        'parts': row['total_parts'],
                        'materials': row['material_count'],
                        'types': row['type_count']
                    },
                    'results': {
                        'body': json.loads(row['body_data'] or '[]'),
                        'thin': json.loads(row['thin_data'] or '[]')
                    }
                } for row in rows]
        except Exception as e:
            print(f"Database error: {e}")
            return []

    def get_history_by_ids(self, job_ids: List[int]) -> List[Dict]:
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                placeholders = ','.join('?' * len(job_ids))
                cursor.execute(f"SELECT * FROM history WHERE id IN ({placeholders})", job_ids)
                rows = cursor.fetchall()

                return [{
                    'id': row['id'],
                    'job_no': row['job_no'],
                    'date': row['date'],
                    'file_name': row['file_name'],
                    'file_path': row['file_path'],
                    'output_path': row['output_path'],
                    'stats': {
                        'parts': row['total_parts'],
                        'materials': row['material_count'],
                        'types': row['type_count']
                    },
                    'results': {
                        'body': json.loads(row['body_data'] or '[]'),
                        'thin': json.loads(row['thin_data'] or '[]')
                    }
                } for row in rows]
        except Exception as e:
            print(f"Database error: {e}")
            return []

    def get_stats(self) -> Dict:
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM stats WHERE id = 1")
                row = cursor.fetchone()

                today = datetime.now().strftime("%Y-%m-%d")
                cursor.execute("""
                    SELECT COUNT(*) as count FROM history WHERE date LIKE ?
                """, (f"{today}%",))
                today_count = cursor.fetchone()['count']
                
                # Malzeme sayısını JsonDataManager'dan al
                try:
                    materials_count = len(JsonDataManager().get_materials())
                except:
                    materials_count = 0

                return {
                    'jobs': row['total_jobs'] if row else 0,
                    'parts': row['total_parts'] if row else 0,
                    'today': today_count,
                    'materials': materials_count
                }
        except Exception as e:
            print(f"Database error: {e}")
            return {'jobs': 0, 'parts': 0, 'today': 0, 'materials': 0}

# ============================================================
# JSON DATA MANAGER
# ============================================================

class JsonDataManager:
    """JSON file manager for settings and materials"""

    def __init__(self):
        self.default_settings = {
            "standart_yukseklik": 720,
            "standart_derinlik": 580,
            "ust_dolap_yukseklik": 720,
            "ust_dolap_derinlik": 330,
            "boy_dolap_yukseklik": 2100,
            "boy_dolap_derinlik": 580,
            "yan_dusumu": 36,
            "raf_genislik_dusumu": 37,
            "raf_derinlik_alt_dolap": 50,
            "raf_derinlik_ust_dolap": 40,
            "sabit_derinlik_dusumu": 23,  # Sabit parça için derinlik düşümü
            "arkalik_dusumu": 18,
            "arkalik_icerde_dusumu": 37,  # Arkalık içerde için düşüm
            "cekmece_alti_dusumu": 60,
            "tolerans": 5,
            "arkalik_max_kalinlik": 8,
            "kanalli_ayir": True,
            "govde_kalinlik": 18,  # Gövde için varsayılan kalınlık
            "cekmece_yan_kalinlik": 16,  # Çekmece yanları için kalınlık
            "arkalik_kalinlik": 8  # Arkalık için kalınlık
        }

    def _read_json(self, name: str, default=None) -> Any:
        if default is None:
            default = {}
        path = Config.get_json_path(name)
        if not path.exists():
            return default
        try:
            with open(str(path), 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return default

    def _write_json(self, name: str, data: Any) -> bool:
        try:
            path = Config.get_json_path(name)
            with open(str(path), 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            print(f"JSON write error: {e}")
            return False

    def get_settings(self) -> Dict:
        settings = self._read_json('settings', {})
        return {**self.default_settings, **settings}

    def save_settings(self, settings: Dict) -> bool:
        return self._write_json('settings', settings)

    def get_default_settings(self) -> Dict:
        return self.default_settings.copy()

    def get_materials(self) -> Dict:
        return self._read_json('materials', {})

    def save_material(self, code: str, thickness: int) -> bool:
        materials = self.get_materials()
        materials[code] = thickness
        return self._write_json('materials', materials)

    def save_all_materials(self, materials: Dict) -> bool:
        return self._write_json('materials', materials)

    def delete_material(self, code: str) -> bool:
        materials = self.get_materials()
        if code in materials:
            del materials[code]
            return self._write_json('materials', materials)
        return False

    def clear_materials(self) -> bool:
        return self._write_json('materials', {})

    # ============================================================
    # LEARNED PARTS - Öğrenilen Parça Tipleri
    # ============================================================
    
    def get_learned_parts(self) -> Dict:
        """Öğrenilen parça tiplerini döndür - {ölçü_malzeme: parça_tipi}"""
        return self._read_json('learned_parts', {})
    
    def save_learned_part(self, boy: int, en: int, malzeme: str, part_type: str) -> bool:
        """Tek bir parça tipini öğren"""
        learned = self.get_learned_parts()
        key = f"{boy}x{en}_{malzeme}"
        learned[key] = part_type
        return self._write_json('learned_parts', learned)
    
    def save_learned_parts_bulk(self, parts: List[Dict]) -> bool:
        """Birden fazla parça tipini öğren"""
        learned = self.get_learned_parts()
        for part in parts:
            key = f"{part['boy']}x{part['en']}_{part['malzeme']}"
            learned[key] = part['partType']
        return self._write_json('learned_parts', learned)
    
    def get_learned_part_type(self, boy: int, en: int, malzeme: str) -> Optional[str]:
        """Öğrenilmiş parça tipini getir"""
        learned = self.get_learned_parts()
        key = f"{boy}x{en}_{malzeme}"
        return learned.get(key)
    
    def clear_learned_parts(self) -> bool:
        """Tüm öğrenilmiş parça tiplerini sil"""
        return self._write_json('learned_parts', {})

# ============================================================
# USER MANAGER
# ============================================================

class UserManager:
    """User authentication manager"""

    def __init__(self):
        self._users_file = Config.get_json_path('users')
        self._ensure_admin()

    def _ensure_admin(self):
        users = self._get_users()
        if 'admin' not in users:
            self._create_user('admin', 'admin123', 'Yönetici', True)

    def _get_users(self) -> Dict:
        try:
            if not self._users_file.exists():
                return {}
            with open(str(self._users_file), 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}

    def _save_users(self, users: Dict) -> bool:
        try:
            with open(str(self._users_file), 'w', encoding='utf-8') as f:
                json.dump(users, f, ensure_ascii=False, indent=4)
            return True
        except:
            return False

    def _hash_password(self, password: str) -> str:
        if BCRYPT_AVAILABLE:
            return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        else:
            return hashlib.sha256(password.encode()).hexdigest()

    def _verify_password(self, password: str, hashed: str) -> bool:
        if BCRYPT_AVAILABLE:
            try:
                return bcrypt.checkpw(password.encode(), hashed.encode())
            except:
                return hashlib.sha256(password.encode()).hexdigest() == hashed
        else:
            return hashlib.sha256(password.encode()).hexdigest() == hashed

    def login(self, username: str, password: str, remember_me: bool = False) -> Dict:
        users = self._get_users()
        if username not in users:
            return {'success': False, 'error': 'Kullanıcı bulunamadı'}

        user = users[username]
        if not self._verify_password(password, user['password']):
            return {'success': False, 'error': 'Şifre hatalı'}

        result = {
            'success': True,
            'user': user.get('display_name', username),
            'is_admin': user.get('is_admin', False)
        }
        
        # Beni Hatırla seçiliyse token oluştur
        if remember_me:
            result['session_token'] = self.generate_session_token(username)
        
        return result

    def _create_user(self, username: str, password: str,
                     display_name: str = '', is_admin: bool = False) -> bool:
        users = self._get_users()
        users[username] = {
            'password': self._hash_password(password),
            'display_name': display_name or username,
            'is_admin': is_admin
        }
        return self._save_users(users)

    def add_user(self, username: str, password: str,
                 display_name: str = '', is_admin: bool = False) -> Dict:
        users = self._get_users()
        if username in users:
            return {'success': False, 'error': 'Kullanıcı zaten mevcut'}

        if self._create_user(username, password, display_name, is_admin):
            return {'success': True}
        return {'success': False, 'error': 'Kullanıcı eklenemedi'}

    def delete_user(self, username: str) -> Dict:
        if username == 'admin':
            return {'success': False, 'error': 'Admin silinemez'}

        users = self._get_users()
        if username in users:
            del users[username]
            if self._save_users(users):
                return {'success': True}
        return {'success': False, 'error': 'Kullanıcı silinemedi'}

    def get_users_list(self) -> List[Dict]:
        users = self._get_users()
        return [{
            'username': k,
            'display_name': v.get('display_name', k),
            'is_admin': v.get('is_admin', False)
        } for k, v in users.items()]
    
    # === Session Token Functions (Beni Hatırla) ===
    def generate_session_token(self, username: str) -> str:
        """Kullanıcı için benzersiz session token oluştur"""
        import secrets
        token = secrets.token_urlsafe(32)
        
        users = self._get_users()
        if username in users:
            users[username]['session_token'] = token
            self._save_users(users)
        
        return token
    
    def verify_session_token(self, username: str, token: str) -> Dict:
        """Token ile kullanıcı doğrula"""
        users = self._get_users()
        if username not in users:
            return {'success': False, 'error': 'Kullanıcı bulunamadı'}
        
        user = users[username]
        stored_token = user.get('session_token')
        
        if not stored_token or stored_token != token:
            return {'success': False, 'error': 'Token geçersiz'}
        
        return {
            'success': True,
            'user': user.get('display_name', username),
            'is_admin': user.get('is_admin', False)
        }
    
    def clear_session_token(self, username: str) -> bool:
        """Çıkış yaparken token'ı temizle"""
        users = self._get_users()
        if username in users:
            users[username]['session_token'] = None
            self._save_users(users)
            # Oturum dosyasını da temizle
            self._clear_session_file()
            return True
        return False
    
    # === Session File (PyWebview localStorage persist etmiyor) ===
    def _get_session_file_path(self) -> Path:
        """Oturum dosyası yolu"""
        return Config.get_data_dir() / 'core_session.json'
    
    def save_session_file(self, username: str, token: str) -> bool:
        """Oturumu dosyaya kaydet"""
        try:
            session_data = {
                'username': username,
                'token': token
            }
            with open(str(self._get_session_file_path()), 'w', encoding='utf-8') as f:
                json.dump(session_data, f, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"Session save error: {e}")
            return False
    
    def load_session_file(self) -> Dict:
        """Kaydedilmiş oturumu oku"""
        try:
            session_path = self._get_session_file_path()
            if not session_path.exists():
                return {'success': False, 'error': 'Kayıtlı oturum yok'}
            
            with open(str(session_path), 'r', encoding='utf-8') as f:
                session_data = json.load(f)
            
            username = session_data.get('username')
            token = session_data.get('token')
            
            if not username or not token:
                return {'success': False, 'error': 'Geçersiz oturum verisi'}
            
            # Token'ı doğrula
            return self.verify_session_token(username, token)
        except Exception as e:
            print(f"Session load error: {e}")
            return {'success': False, 'error': str(e)}
    
    def _clear_session_file(self) -> bool:
        """Oturum dosyasını sil"""
        try:
            session_path = self._get_session_file_path()
            if session_path.exists():
                session_path.unlink()
            return True
        except:
            return False



# ============================================================
# EXCEL ANALYZER - BELGEYE UYGUN VERSİYON
# ============================================================

class ExcelAnalyzer:
    """
    Excel file analyzer and processor
    
    BELGE KURALLARI:
    - E sütunu: POZ numarası
    - F sütunu: Modül adı (Alt dolap 60 cm, Üst dolap 60 cm, vb.)
    - H sütunu: Adet
    - I sütunu: Ölçü 1
    - J sütunu: Ölçü 2
    - K sütunu: Kanallı bilgisi
    - L sütunu: Malzeme adı
    
    KALINLIK KURALLARI:
    - 18mm: Gövde (yan, alt-üst, raf, kayıt/kuşak)
    - 16mm: Çekmece yanları
    - 8mm: Arkalık
    
    PARÇA FORMÜLLER:
    - YAN: Yükseklik x Derinlik (720x580 alt, 720x330 üst)
    - ALT-ÜST: (Genişlik-36) x (Derinlik-1)
    - RAF Alt: (Genişlik-37) x (Derinlik-50)
    - RAF Üst: (Genişlik-37) x (Derinlik-40)
    - ARKALIK: (Yükseklik-18) x (Genişlik-18)
    - KAYIT/KUŞAK: Bir kenar 100 veya 120mm civarı
    """

    def __init__(self, data_manager: JsonDataManager, custom_modules: Dict = None, cabinet_settings: Dict = None):
        self.data_manager = data_manager
        self.custom_modules = custom_modules or {}
        # Toplu dolap tipi ayarları
        self.cabinet_settings = cabinet_settings or {
            'alt': {'yukseklik': 720, 'derinlik': 580},
            'ust': {'yukseklik': 720, 'derinlik': 330},
            'boy': {'yukseklik': 2100, 'derinlik': 580}
        }
    
    def _get_custom_module_for_poz(self, poz: str) -> Dict:
        """POZ için özel modül ayarlarını döndür"""
        return self.custom_modules.get(poz)

    def check_file(self, file_path: str) -> Dict:
        """Dosyayı kontrol et ve bilinmeyen malzemeleri bul"""
        if not PANDAS_AVAILABLE:
            return {'success': False, 'error': 'pandas kütüphanesi yüklü değil!'}

        try:
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)

            # Sütun kontrolü - Esnek sütun isimleri
            # Önce standart isimleri dene, yoksa indeks bazlı erişim
            col_mapping = self._get_column_mapping(df)
            
            if not col_mapping:
                return {'success': False, 'error': 'Excel formatı tanınamadı. Gerekli sütunlar bulunamadı.'}

            # Job number - Info16'dan al
            job_no = None
            if '§542 Info16' in df.columns:
                info16_values = df['§542 Info16'].dropna().unique()
                if len(info16_values) > 0:
                    job_no = str(info16_values[0]).strip()

            # Malzemeleri al (L sütunu veya Malzeme Kodu)
            malzeme_col = col_mapping.get('malzeme')
            if malzeme_col and malzeme_col in df.columns:
                materials = df[malzeme_col].dropna().unique().tolist()
                materials = [str(m).strip() for m in materials if str(m).strip() and str(m).strip().lower() != 'nan']
            else:
                materials = []

            known = self.data_manager.get_materials()
            unknown = [m for m in materials if m not in known]

            return {
                'success': True,
                'materials': materials,
                'unknown': unknown,
                'row_count': len(df),
                'material_count': len(materials),
                'job_no': job_no
            }
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def _get_column_mapping(self, df) -> Dict:
        """Excel sütunlarını belirle - esnek mapping"""
        mapping = {}
        columns = df.columns.tolist()
        
        # Debug: Sütunları yazdır
        print(f"Excel columns: {columns}")
        
        # Öncelikli olarak bilinen sütun isimlerini ara
        for col in columns:
            col_str = str(col)
            col_lower = col_str.lower().strip()
            
            # POZ (Info4) - hem §542 hem #8542 destekle
            if col_lower in ['poz', 'poz no', 'poz numarası']:
                mapping['poz'] = col
            elif 'Info4' in col_str:
                mapping['poz'] = col
            
            # Modül adı (Info5) - GENİŞLİK BURADAN ALINACAK
            if col_lower in ['modül', 'modul', 'modül adı', 'modul adi']:
                mapping['modul'] = col
            elif 'Info5' in col_str:
                mapping['modul'] = col
            
            # Modül kodu (Info3) - Genişlik kontrolü için
            if 'Info3' in col_str:
                mapping['modul_kodu'] = col
            
            # Adet/Sipariş (H sütunu) - #8542 formatında "Sipariş" veya "Info2" olabilir
            if col_lower in ['adet', 'miktar', 'qty', 'quantity', 'sipariş', 'siparis', 'siparış']:
                mapping['adet'] = col
            # #8542 formatında Sipariş sütunu header'da "Sipariş" veya başka bir şey olabilir
            # Sütun adında "Sipariş" veya "sipari" varsa
            elif 'sipari' in col_lower or 'sipariş' in col_lower:
                mapping['adet'] = col
            
            # Ölçü 1 (I sütunu) - Uzunluk (parça boyu)
            if col_lower in ['uzunluk', 'boy', 'length', 'ölçü 1', 'olcu 1']:
                mapping['olcu1'] = col
            
            # Ölçü 2 (J sütunu) - Genişlik (parça eni) - SÜTUN VARSA
            if col_lower in ['genişlik', 'genislik', 'en', 'width', 'ölçü 2', 'olcu 2']:
                mapping['olcu2'] = col
            
            # Kanallı (Info1) - hem §542 hem #8542 destekle
            if col_lower in ['kanallı', 'kanalli', 'kanal']:
                mapping['kanalli'] = col
            elif 'Info1' in col_str:
                mapping['kanalli'] = col
            
            # Malzeme (L sütunu)
            if col_lower in ['malzeme', 'malzeme kodu', 'material', 'malzeme adı']:
                mapping['malzeme'] = col
            
            # Info16 (İş numarası)
            if 'Info16' in col_str:
                mapping['info16'] = col
        
        # Alternatif isimler
        if 'olcu1' not in mapping:
            for col in columns:
                if 'Uzunluk' in str(col):
                    mapping['olcu1'] = col
                    break
        
        # Genişlik sütunu yoksa - parça ölçüsü olarak Uzunluk'tan sonraki sütunu dene
        if 'olcu2' not in mapping and 'olcu1' in mapping:
            olcu1_idx = columns.index(mapping['olcu1'])
            # Bir sonraki sütuna bak
            for i in range(olcu1_idx + 1, min(olcu1_idx + 3, len(columns))):
                col = columns[i]
                col_str = str(col)
                # Info veya Malzeme değilse ve sayısal içerik varsa
                if 'Info' not in col_str and 'Malzeme' not in col_str and 'Kesim' not in col_str:
                    mapping['olcu2'] = col
                    print(f"Parça eni sütunu otomatik bulundu: {col}")
                    break
        
        # ADET sütunu bulunamadıysa - "Sipariş" kelimesini içeren sütunu ara
        if 'adet' not in mapping:
            for col in columns:
                col_str = str(col)
                if 'Sipariş' in col_str or 'Siparis' in col_str:
                    mapping['adet'] = col
                    print(f"Sipariş sütunu bulundu: {col}")
                    break
        
        if 'malzeme' not in mapping:
            for col in columns:
                if 'Malzeme' in str(col):
                    mapping['malzeme'] = col
                    break
        
        print(f"Column mapping: {mapping}")
        return mapping

    def analyze_and_export(self, file_path: str, output_path: str, custom_depths: Dict = None) -> Dict:
        """
        Excel dosyasını analiz et ve kesim listesi oluştur.
        
        BELGE KURALLARI UYGULANIR:
        - Kalınlıklar: 18mm (gövde), 16mm (çekmece yan), 8mm (arkalık)
        - Parça formülleri belgeye göre
        """
        if not PANDAS_AVAILABLE:
            return {'success': False, 'error': 'pandas kütüphanesi yüklü değil!'}

        try:
            # Dosyayı oku
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)

            settings = self.data_manager.get_settings()
            materials_db = self.data_manager.get_materials()
            
            # Sütun mapping
            col_map = self._get_column_mapping(df)

            # Job number - Info16'dan al
            job_no = None
            if '§542 Info16' in df.columns:
                info16_values = df['§542 Info16'].dropna().unique()
                if len(info16_values) > 0:
                    job_no = str(info16_values[0]).strip()

            # ============================================================
            # STANDART ÖLÇÜLER (Belgeden)
            # ============================================================
            ALT_DOLAP_YUKSEKLIK = settings.get('standart_yukseklik', 720)
            ALT_DOLAP_DERINLIK = settings.get('standart_derinlik', 580)
            UST_DOLAP_YUKSEKLIK = settings.get('ust_dolap_yukseklik', 720)
            UST_DOLAP_DERINLIK = settings.get('ust_dolap_derinlik', 330)
            BOY_DOLAP_YUKSEKLIK = settings.get('boy_dolap_yukseklik', 2100)
            BOY_DOLAP_DERINLIK = settings.get('boy_dolap_derinlik', 580)
            
            # Düşüm değerleri (Belgeden)
            YAN_DUSUM = 36  # Alt-üst için genişlikten düşülen
            ALT_UST_DERINLIK_DUSUM = 1  # Alt-üst için derinlikten düşülen
            RAF_GENISLIK_DUSUM = 37  # Raf için genişlikten düşülen
            RAF_DERINLIK_ALT = 50  # Alt dolap rafı için derinlikten düşülen
            RAF_DERINLIK_UST = 40  # Üst dolap rafı için derinlikten düşülen
            ARKALIK_DUSUM = 18  # Arkalık için her iki kenardan düşülen
            
            # Sabit kalınlıklar (Belgeden)
            GOVDE_KALINLIK = 18
            CEKMECE_YAN_KALINLIK = 16
            ARKALIK_KALINLIK = 8
            
            TOLERANS = settings.get('tolerans', 10)  # 10mm tolerans - daha esnek eşleşme

            # ============================================================
            # VERİ HAZIRLAMA
            # ============================================================
            
            # Ölçüleri al
            olcu1_col = col_map.get('olcu1', 'Uzunluk')
            olcu2_col = col_map.get('olcu2', 'Genişlik')
            malzeme_col = col_map.get('malzeme', 'Malzeme Kodu')
            modul_col = col_map.get('modul', '§542 Info5')
            poz_col = col_map.get('poz', '§542 Info4')
            kanalli_col = col_map.get('kanalli', '§542 Info1')
            
            # DataFrame'e standart sütunlar ekle
            df['OLCU1'] = pd.to_numeric(df[olcu1_col], errors='coerce').fillna(0).astype(int)
            df['OLCU2'] = pd.to_numeric(df[olcu2_col], errors='coerce').fillna(0).astype(int)
            df['MALZEME'] = df[malzeme_col].fillna('').astype(str).str.strip() if malzeme_col in df.columns else ''
            
            # ADET sütununu Excel'den oku (Sipariş sütunu)
            adet_col = col_map.get('adet')
            if adet_col and adet_col in df.columns:
                df['ADET'] = pd.to_numeric(df[adet_col], errors='coerce').fillna(1).astype(int)
                print(f"ADET sütunu bulundu: {adet_col}")
            else:
                df['ADET'] = 1
                print("ADET sütunu bulunamadı, varsayılan 1 kullanılıyor")
            
            # POZ ve Modül bilgisi
            if modul_col in df.columns:
                df['MODUL_ADI'] = df[modul_col].fillna('').astype(str)
            else:
                df['MODUL_ADI'] = ''
                
            if poz_col in df.columns:
                df['POZ'] = df[poz_col].fillna('').astype(str)
            else:
                df['POZ'] = ''

            # ============================================================
            # MODÜL TİPİ VE GENİŞLİK BELİRLEME
            # ============================================================
            
            def parse_module_info(modul_adi: str) -> tuple:
                """
                Modül adından tip ve genişlik çıkar.
                Örnek: "Alt dolap 60 cm" -> ('ALT', 600)
                """
                modul_adi = str(modul_adi).lower().strip()
                
                # Tip belirleme
                if 'üst' in modul_adi or 'ust' in modul_adi:
                    tip = 'ÜST'
                elif 'boy' in modul_adi:
                    tip = 'BOY'
                else:
                    tip = 'ALT'
                
                # Genişlik belirleme (cm'den mm'ye)
                genislik = None
                cm_match = re.search(r'(\d+)\s*cm', modul_adi)
                if cm_match:
                    genislik = int(cm_match.group(1)) * 10  # cm -> mm
                
                return tip, genislik
            
            # Her satır için modül bilgisi
            df['MODUL_TIP'] = ''
            df['MODUL_GENISLIK'] = 0
            
            for idx, row in df.iterrows():
                tip, genislik = parse_module_info(row['MODUL_ADI'])
                df.at[idx, 'MODUL_TIP'] = tip
                df.at[idx, 'MODUL_GENISLIK'] = genislik if genislik else 600  # Varsayılan 600mm

            # ============================================================
            # KANALLI TESPİTİ
            # ============================================================
            
            def is_kanalli(row) -> bool:
                """K sütunundaki veya Info1'deki kanallı bilgisini kontrol et"""
                if kanalli_col not in df.columns:
                    return False
                
                kanalli_val = str(row.get(kanalli_col, '')).upper().strip()
                
                # True/False kontrolü
                if kanalli_val in ['TRUE', 'EVET', 'YES', '1', 'VAR']:
                    return True
                
                # Pattern kontrolü: SOL_13+9 veya SAĞ_xxx+xxx
                if re.search(r'(SOL|SAĞ|SAG)_\d+\+\d+', kanalli_val):
                    return True
                
                return False

            # ============================================================
            # ÖĞRENİLMİŞ PARÇALAR
            # ============================================================
            learned_parts = self.data_manager.get_learned_parts()

            # ============================================================
            # PARÇA TİPİ BELİRLEME - BELGE KURALLARINA GÖRE
            # ============================================================
            
            def determine_part_type(row) -> tuple:
                """
                Parça tipini ve kalınlığını ÖLÇÜLERDEN TERSİNE HESAPLAMA ile belirle.
                
                MANTIK:
                0. Önce öğrenilmiş parçalara bak
                1. Ölçülere bakarak hangi formüle uyduğunu bul
                2. Formülden modül genişliğini ve tipini çıkar
                
                FORMÜLLER (Belgeden):
                - YAN: 720 x 580 (Alt), 720 x 330 (Üst), 2100 x 580 (Boy)
                - ALT-ÜST: (Genişlik-36) x (Derinlik-1) → 564x579, 664x579, vb.
                - RAF Alt: (Genişlik-37) x (Derinlik-50) → 563x530, 663x530, vb.
                - RAF Üst: (Genişlik-37) x (Derinlik-40) → 563x290, 663x290, vb.
                - ARKALIK 8mm: (Yükseklik-18) x (Genişlik-18)
                - KAYIT/KUŞAK: Bir kenar 80-140mm civarı
                
                Returns: (parça_tipi, kalınlık, boy, en)
                """
                olcu1 = row['OLCU1']
                olcu2 = row['OLCU2']
                kanalli = is_kanalli(row)
                
                # Malzeme kalınlığını veritabanından al
                malzeme = row['MALZEME']
                db_kalinlik = materials_db.get(malzeme, GOVDE_KALINLIK)
                
                # Ölçüleri sırala - büyük olan BOY, küçük olan EN
                boy = max(olcu1, olcu2)
                en = min(olcu1, olcu2)
                
                # ============================================================
                # 0. ÖĞRENİLMİŞ PARÇA KONTROLÜ - En önce bak!
                # ============================================================
                learned_key = f"{boy}x{en}_{malzeme}"
                if learned_key in learned_parts:
                    learned_type = learned_parts[learned_key]
                    # Kalınlık HER ZAMAN malzemenin veritabanındaki kalınlığı!
                    return learned_type, db_kalinlik, boy, en
                
                # ============================================================
                # 1. MALZEME KALINLIĞINA GÖRE TABLO BELİRLE
                # 8mm ve altı → İnce (Arkalık) tablosu
                # 8mm üstü → Gövde tablosu
                # ============================================================
                
                # Arkalık (8mm ve altı) - Arkalık tiplerini kontrol et
                if db_kalinlik <= ARKALIK_KALINLIK:
                    return 'ARKALIK', db_kalinlik, boy, en
                
                # Çekmece yanı (16mm) - veritabanında 16mm olarak işaretli
                if db_kalinlik == CEKMECE_YAN_KALINLIK:
                    return 'ÇEKMECE YANI', db_kalinlik, boy, en
                
                # ============================================================
                # 2. YAN KONTROLÜ - Standart yükseklik ve derinlik
                # Alt dolap: 720 x 580
                # Üst dolap: 720 x 330
                # Boy dolap: 2100 x 580
                # ============================================================
                
                # Üst dolap yanı: 720 x 330
                if (abs(boy - 720) <= TOLERANS and abs(en - 330) <= TOLERANS):
                    if kanalli:
                        return 'YAN (KANALLI)', db_kalinlik, boy, en
                    return 'YAN', db_kalinlik, boy, en
                
                # Alt dolap yanı: 720 x 580
                if (abs(boy - 720) <= TOLERANS and abs(en - 580) <= TOLERANS):
                    if kanalli:
                        return 'YAN (KANALLI)', db_kalinlik, boy, en
                    return 'YAN', db_kalinlik, boy, en
                
                # Boy dolap yanı: 2100 x 580
                if (abs(boy - 2100) <= TOLERANS and abs(en - 580) <= TOLERANS):
                    if kanalli:
                        return 'YAN (KANALLI)', db_kalinlik, boy, en
                    return 'YAN', db_kalinlik, boy, en
                
                # Özel derinlikli yan (custom_depths kullanılıyorsa)
                if abs(boy - 720) <= TOLERANS or abs(boy - 2100) <= TOLERANS:
                    # Yükseklik standart, derinlik kontrol
                    if 300 <= en <= 600:  # Makul derinlik aralığı
                        if kanalli:
                            return 'YAN (KANALLI)', db_kalinlik, boy, en
                        # Sadece standart derinliklere yakınsa YAN de
                        if abs(en - 580) <= TOLERANS or abs(en - 330) <= TOLERANS:
                            return 'YAN', db_kalinlik, boy, en
                
                # ============================================================
                # 3. ALT-ÜST KONTROLÜ - (Genişlik-36) x (Derinlik-1)
                # Tersine: Genişlik = BOY + 36
                # EN değeri: 579 (580-1), 329 (330-1), vb.
                # ============================================================
                
                # Alt dolap alt-üst: EN ≈ 579 (580-1)
                if abs(en - 579) <= TOLERANS:
                    # Modül genişliği = BOY + 36
                    modul_gen = boy + 36
                    if modul_gen % 50 <= TOLERANS or (50 - modul_gen % 50) <= TOLERANS:  # 50'nin katı (500, 550, 600, vb.)
                        if kanalli:
                            return 'ALT-ÜST (KANALLI)', db_kalinlik, boy, en
                        return 'ALT-ÜST', db_kalinlik, boy, en
                
                # Üst dolap alt-üst: EN ≈ 329 (330-1)
                if abs(en - 329) <= TOLERANS:
                    modul_gen = boy + 36
                    if modul_gen % 50 <= TOLERANS or (50 - modul_gen % 50) <= TOLERANS:
                        if kanalli:
                            return 'ALT-ÜST (KANALLI)', db_kalinlik, boy, en
                        return 'ALT-ÜST', db_kalinlik, boy, en
                
                # Genel alt-üst kontrolü - EN değeri derinlik-1 olmalı
                # 580-1=579, 330-1=329, veya özel derinlikler
                possible_depths = [579, 329, 549, 529, 559]  # Yaygın derinlik-1 değerleri
                for depth in possible_depths:
                    if abs(en - depth) <= TOLERANS:
                        if kanalli:
                            return 'ALT-ÜST (KANALLI)', db_kalinlik, boy, en
                        return 'ALT-ÜST', db_kalinlik, boy, en
                
                # ============================================================
                # 4. RAF KONTROLÜ - (Genişlik-37) x (Derinlik-50 veya 40)
                # RAF ASLA KANALLI OLMAZ!
                # Alt dolap raf: EN ≈ 530 (580-50)
                # Üst dolap raf: EN ≈ 290 (330-40)
                # ============================================================
                
                if not kanalli:  # RAF ASLA KANALLI OLMAZ
                    # Alt dolap rafı: EN ≈ 530
                    if abs(en - 530) <= TOLERANS:
                        modul_gen = boy + 37
                        return 'RAF', db_kalinlik, boy, en
                    
                    # Üst dolap rafı: EN ≈ 290
                    if abs(en - 290) <= TOLERANS:
                        modul_gen = boy + 37
                        return 'RAF (ÜST)', db_kalinlik, boy, en
                    
                    # Diğer olası raf derinlikleri (özel modüller için)
                    # Derinlik-50 veya Derinlik-40 formülüne uyan değerler
                    raf_depths_alt = [530, 520, 510, 500, 480, 450]  # Derinlik-50
                    raf_depths_ust = [290, 280, 270, 260]  # Derinlik-40
                    
                    for raf_en in raf_depths_alt:
                        if abs(en - raf_en) <= TOLERANS:
                            return 'RAF', db_kalinlik, boy, en
                    
                    for raf_en in raf_depths_ust:
                        if abs(en - raf_en) <= TOLERANS:
                            return 'RAF (ÜST)', db_kalinlik, boy, en
                
                # ============================================================
                # 5. KAYIT/KUŞAK KONTROLÜ - Bir kenar 80-140mm civarı
                # ============================================================
                
                if 80 <= en <= 140:
                    return 'KAYIT/KUŞAK', db_kalinlik, boy, en
                
                if 80 <= boy <= 140:
                    return 'KAYIT/KUŞAK', db_kalinlik, boy, en
                
                # ============================================================
                # 6. DİĞER - Hiçbir formüle uymayan
                # Kalınlık her zaman malzemenin kalınlığı!
                # ============================================================
                
                return 'DİĞER', db_kalinlik, boy, en

            # ============================================================
            # TÜM PARÇALARI İŞLE
            # ============================================================
            
            results = []
            for idx, row in df.iterrows():
                parca_tipi, kalinlik, boy, en = determine_part_type(row)
                
                results.append({
                    'POZ': row['POZ'],
                    'MODÜL': row['MODUL_ADI'],
                    'MODÜL TİPİ': row['MODUL_TIP'],
                    'KALINLIK': kalinlik,
                    'MALZEME': row['MALZEME'],
                    'BOY': boy,
                    'EN': en,
                    'PARÇA TİPİ': parca_tipi,
                    'ADET': row['ADET']
                })
            
            result_df = pd.DataFrame(results)
            
            # ============================================================
            # GRUPLAMA VE TOPLAMA
            # ============================================================
            
            group_cols = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ']
            summary = result_df.groupby(group_cols).agg({'ADET': 'sum'}).reset_index()
            
            # Kalınlığa göre 3 tabloya ayır (18mm, 16mm, 8mm)
            df_18mm = summary[summary['KALINLIK'] == 18].copy()
            df_16mm = summary[summary['KALINLIK'] == 16].copy()
            df_8mm = summary[summary['KALINLIK'] <= 8].copy()
            
            # PARÇA TİPİ'ne göre sırala - aynı tipler alt alta gelsin
            part_type_order = [
                'YAN', 'YAN (K)',
                'ALT-ÜST', 'ALT-ÜST (K)',
                'SABİT', 'SABİT (K)',
                'RAF', 'RAF (K)',
                'RAF (ÜST)', 'RAF (ÜST) (K)',
                'KAYIT/KUŞAK', 'KAYIT/KUŞAK (K)',
                'ÇEKMECE YANI', 'ÇEKMECE YANI (K)',
                'ARKALIK', 'ARKALIK (K)',
                'ARKALIK (İÇERDE)', 'ARKALIK (İÇERDE) (K)',
                'DİĞER', 'DİĞER (K)'
            ]
            
            def sort_by_part_type(df):
                if df.empty or 'PARÇA TİPİ' not in df.columns:
                    return df
                df['_sort_order'] = df['PARÇA TİPİ'].apply(
                    lambda x: part_type_order.index(x) if x in part_type_order else 999
                )
                df = df.sort_values(by=['_sort_order', 'MALZEME', 'BOY', 'EN'])
                df = df.drop(columns=['_sort_order'])
                return df
            
            df_18mm = sort_by_part_type(df_18mm)
            df_16mm = sort_by_part_type(df_16mm)
            df_8mm = sort_by_part_type(df_8mm)


            # ============================================================
            # EXCEL'E YAZI - TEK SHEET'TE 3 TABLO YAN YANA
            # ============================================================
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Kesim Listesi'
            
            # Stil tanımları - 3 farklı renk
            header_font = Font(bold=True, color='FFFFFF')
            header_fill_18mm = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Mavi
            header_fill_16mm = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')  # Mor
            header_fill_8mm = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')   # Yeşil
            title_font = Font(bold=True, size=14)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sütun sıralaması
            column_order = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ', 'ADET']
            
            # DataFrame'leri sütun sırasına göre düzenle
            for df in [df_18mm, df_16mm, df_8mm]:
                if not df.empty:
                    cols = [c for c in column_order if c in df.columns]
                    df = df[cols]
            
            # Sabit sütun sayısı
            cols_count = 6
            table_gap = 1  # Tablolar arası boşluk
            
            def write_table(ws, start_col, df, title, header_fill):
                """Tek bir tabloyu belirtilen sütundan başlayarak yaz"""
                current_row = 1
                
                # Zebra efekti için gri renk
                gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # Başlık - ORTALI
                ws.cell(row=current_row, column=start_col, value=title)
                ws.cell(row=current_row, column=start_col).font = title_font
                ws.cell(row=current_row, column=start_col).alignment = Alignment(horizontal='center')
                if cols_count > 1:
                    ws.merge_cells(start_row=current_row, start_column=start_col, 
                                  end_row=current_row, end_column=start_col + cols_count - 1)
                
                current_row += 1
                
                # Header satırı
                header_names = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ', 'ADET']
                for col_idx, col_name in enumerate(header_names):
                    cell = ws.cell(row=current_row, column=start_col + col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                current_row += 1
                
                # Veri satırları - ZEBRA EFEKTİ (beyaz/gri)
                if not df.empty:
                    # DataFrame sütunlarını sırala
                    df_ordered = df.copy()
                    existing_cols = [c for c in column_order if c in df_ordered.columns]
                    df_ordered = df_ordered[existing_cols]
                    
                    row_index = 0
                    for row_data in df_ordered.values:
                        for col_idx, value in enumerate(row_data):
                            cell = ws.cell(row=current_row, column=start_col + col_idx, value=value)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # Çift satırlar gri (0, 2, 4...), tek satırlar beyaz
                            if row_index % 2 == 1:
                                cell.fill = gray_fill
                        current_row += 1
                        row_index += 1
                else:
                    ws.cell(row=current_row, column=start_col, value='(Veri yok)')
                    current_row += 1
                
                return current_row
            
            # ========== 1. TABLO: 18mm PARÇALAR (Mavi) ==========
            col_18mm_start = 1
            write_table(worksheet, col_18mm_start, df_18mm, '18mm Parçalar', header_fill_18mm)
            
            # ========== 2. TABLO: 16mm PARÇALAR (Mor) ==========
            col_16mm_start = col_18mm_start + cols_count + table_gap
            write_table(worksheet, col_16mm_start, df_16mm, '16mm Parçalar', header_fill_16mm)
            
            # ========== 3. TABLO: 8mm PARÇALAR (Yeşil) ==========
            col_8mm_start = col_16mm_start + cols_count + table_gap
            write_table(worksheet, col_8mm_start, df_8mm, '8mm Parçalar', header_fill_8mm)
            
            # Sütun genişliklerini ayarla
            # Sütun sırası: KALINLIK, MALZEME, BOY, EN, PARÇA TİPİ, ADET
            # 18mm ve 16mm için genişlikler
            widths_normal = [9, None, 9.5, 9.5, 14.2, 9]  # None = otomatik
            # 8mm için genişlikler (PARÇA TİPİ farklı)
            widths_8mm = [9, None, 9.5, 9.5, 15.7, 9]
            
            def set_column_widths(start_col, df, widths):
                for i, width in enumerate(widths):
                    col_letter = get_column_letter(start_col + i)
                    if width is None:
                        # Otomatik: içeriğe göre ayarla
                        max_len = len('MALZEME')  # Header uzunluğu
                        if not df.empty and 'MALZEME' in df.columns:
                            max_content = df['MALZEME'].astype(str).str.len().max()
                            max_len = max(max_len, max_content)
                        worksheet.column_dimensions[col_letter].width = max_len + 2
                    else:
                        worksheet.column_dimensions[col_letter].width = width
            
            set_column_widths(col_18mm_start, df_18mm, widths_normal)
            set_column_widths(col_16mm_start, df_16mm, widths_normal)
            set_column_widths(col_8mm_start, df_8mm, widths_8mm)
            
            # İstatistikler
            govde_18 = int(df_18mm['ADET'].sum()) if not df_18mm.empty else 0
            cekmece_16 = int(df_16mm['ADET'].sum()) if not df_16mm.empty else 0
            arkalik_8 = int(df_8mm['ADET'].sum()) if not df_8mm.empty else 0
            
            # Excel dosyasını kaydet
            workbook.save(output_path)

            # ============================================================
            # SONUÇ DÖNDÜR
            # ============================================================
            
            return {
                'success': True,
                'output_path': output_path,
                'job_no': job_no,
                'total_parts': int(summary['ADET'].sum()),
                'material_count': len(result_df['MALZEME'].unique()),
                'type_count': len(summary['PARÇA TİPİ'].unique()),
                'thickness_summary': {
                    'govde_18': govde_18,
                    'cekmece_16': cekmece_16,
                    'arkalik_8': arkalik_8
                },
                'body': df_18mm.to_dict('records') if not df_18mm.empty else [],
                'thin': df_8mm.to_dict('records') if not df_8mm.empty else [],
                'df_16mm': df_16mm.to_dict('records') if not df_16mm.empty else []
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def analyze_only(self, file_path: str, custom_depths: Dict = None) -> Dict:
        """
        Excel dosyasını analiz et ama Excel'e kaydetme.
        Sonuçları düzenleme için UI'a döndür.
        
        ÖNEMLİ: Modül genişliği modül adından alınır (örn: "Alt dolap 60 cm" → 600mm)
        """
        if not PANDAS_AVAILABLE:
            return {'success': False, 'error': 'pandas kütüphanesi yüklü değil!'}

        try:
            # Dosyayı oku
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)

            settings = self.data_manager.get_settings()
            materials_db = self.data_manager.get_materials()
            learned_parts = self.data_manager.get_learned_parts()
            
            col_map = self._get_column_mapping(df)

            # Job number - Info16'dan al
            job_no = None
            info16_col = col_map.get('info16')
            if info16_col and info16_col in df.columns:
                info16_values = df[info16_col].dropna().unique()
                if len(info16_values) > 0:
                    job_no = str(info16_values[0]).strip()
            
            # Alternatif Info16 arama
            if not job_no:
                for col in df.columns:
                    if 'Info16' in str(col):
                        info16_values = df[col].dropna().unique()
                        if len(info16_values) > 0:
                            job_no = str(info16_values[0]).strip()
                            break

            # Sabitler
            GOVDE_KALINLIK = 18
            CEKMECE_YAN_KALINLIK = 16
            ARKALIK_KALINLIK = 8
            TOLERANS = settings.get('tolerans', 10)

            # Sütunları hazırla
            olcu1_col = col_map.get('olcu1', 'Uzunluk')
            malzeme_col = col_map.get('malzeme', 'Malzeme Kodu')
            kanalli_col = col_map.get('kanalli')
            modul_col = col_map.get('modul')
            poz_col = col_map.get('poz')  # POZ sütunu - özel modül ayarları için
            
            # Ölçü sütunlarını kontrol et
            if olcu1_col not in df.columns:
                # Uzunluk sütununu bul
                for col in df.columns:
                    if 'Uzunluk' in str(col):
                        olcu1_col = col
                        break
            
            # İkinci ölçü sütunu (parça eni)
            olcu2_col = col_map.get('olcu2')
            if not olcu2_col or olcu2_col not in df.columns:
                # Uzunluk'tan sonraki sütunu dene
                columns = df.columns.tolist()
                if olcu1_col in columns:
                    olcu1_idx = columns.index(olcu1_col)
                    for i in range(olcu1_idx + 1, min(olcu1_idx + 3, len(columns))):
                        col = columns[i]
                        if 'Info' not in str(col) and 'Malzeme' not in str(col):
                            olcu2_col = col
                            break
            
            print(f"Ölçü sütunları: olcu1={olcu1_col}, olcu2={olcu2_col}")
            
            df['OLCU1'] = pd.to_numeric(df[olcu1_col], errors='coerce').fillna(0).astype(int) if olcu1_col in df.columns else 0
            df['OLCU2'] = pd.to_numeric(df[olcu2_col], errors='coerce').fillna(0).astype(int) if olcu2_col and olcu2_col in df.columns else 0
            df['MALZEME'] = df[malzeme_col].fillna('').astype(str).str.strip() if malzeme_col in df.columns else ''
            
            # ADET sütununu Excel'den oku (Sipariş sütunu)
            adet_col = col_map.get('adet')
            if adet_col and adet_col in df.columns:
                df['ADET'] = pd.to_numeric(df[adet_col], errors='coerce').fillna(1).astype(int)
                print(f"ADET sütunu bulundu: {adet_col}")
            else:
                df['ADET'] = 1
                print("ADET sütunu bulunamadı, varsayılan 1 kullanılıyor")

            # Modül adından genişlik çıkarma fonksiyonu
            def get_modul_genislik(modul_adi: str) -> int:
                """Modül adından genişlik çıkar (cm → mm)"""
                if not modul_adi or pd.isna(modul_adi):
                    return None
                modul_adi = str(modul_adi).lower()
                cm_match = re.search(r'(\d+)\s*cm', modul_adi)
                if cm_match:
                    return int(cm_match.group(1)) * 10  # cm → mm
                return None

            # Kanallı kontrol fonksiyonu
            def check_kanalli(row) -> bool:
                if not kanalli_col or kanalli_col not in df.columns:
                    return False
                kanalli_val = str(row.get(kanalli_col, '')).upper().strip()
                if kanalli_val in ['TRUE', 'EVET', 'YES', '1', 'VAR']:
                    return True
                if re.search(r'(SOL|SAĞ|SAG)_\d+\+\d+', kanalli_val):
                    return True
                return False

            def determine_part_type_with_module(row) -> tuple:
                """
                Parça tipini belirle - modül genişliğini ve özel modül ayarlarını dikkate alarak.
                
                KRİTİK: Kalınlık her zaman malzemenin veritabanındaki kalınlığından (db_kalinlik) alınır!
                Parça tipi ölçülere göre belirlenir ama kalınlık malzemeye göre belirlenir.
                
                Returns: (parça_tipi, kalınlık, boy, en, kanalli)
                """
                olcu1, olcu2 = row['OLCU1'], row['OLCU2']
                malzeme = row['MALZEME']
                kanalli = check_kanalli(row)
                db_kalinlik = materials_db.get(malzeme, GOVDE_KALINLIK)
                
                boy = max(olcu1, olcu2)
                en = min(olcu1, olcu2)
                
                # POZ'u al ve özel modül ayarlarını kontrol et
                poz = str(row.get(poz_col, '')).strip() if poz_col and poz_col in df.columns else ''
                custom_module = self._get_custom_module_for_poz(poz) if poz else None
                
                # Modül adını al ve dolap tipini tespit et
                modul_adi = row.get(modul_col, '') if modul_col and modul_col in df.columns else ''
                modul_adi_str = str(modul_adi).lower() if modul_adi else ''
                
                # Dolap tipini belirle: üst, alt, boy
                if 'üst' in modul_adi_str or 'ust' in modul_adi_str:
                    dolap_tipi = 'ust'
                elif 'boy' in modul_adi_str:
                    dolap_tipi = 'boy'
                else:
                    dolap_tipi = 'alt'
                
                # Modül genişliğini al - önce özel ayarlardan, yoksa modül adından
                if custom_module and custom_module.get('genislik'):
                    modul_gen = custom_module.get('genislik')
                else:
                    modul_gen = get_modul_genislik(modul_adi)
                
                # ============================================================
                # ÖLÇÜLERİ BELİRLE - ÖNCELİK SIRASI:
                # 1. POZ-spesifik özel modül ayarları (custom_module)
                # 2. Toplu dolap tipi ayarları (cabinet_settings)
                # 3. Varsayılan değerler
                # ============================================================
                
                # Önce POZ-spesifik ayarları kontrol et
                if custom_module:
                    custom_yukseklik = custom_module.get('yukseklik')
                    custom_derinlik = custom_module.get('derinlik')
                else:
                    custom_yukseklik = None
                    custom_derinlik = None
                
                # Toplu dolap ayarlarından değerleri al
                cabinet = self.cabinet_settings.get(dolap_tipi, {'yukseklik': 720, 'derinlik': 580})
                
                # Final yükseklik ve derinlik - POZ ayarı öncelikli, yoksa cabinet ayarı
                yukseklik = custom_yukseklik if custom_yukseklik else cabinet.get('yukseklik', 720)
                derinlik = custom_derinlik if custom_derinlik else cabinet.get('derinlik', 580)
                
                # Öğrenilmiş parça kontrolü (kanallı dahil key)
                learned_key = f"{boy}x{en}_{malzeme}"
                if learned_key in learned_parts:
                    learned_type = learned_parts[learned_key]
                    # Öğrenilmiş tip ne olursa olsun, kalınlık malzemenin kalınlığı!
                    return learned_type, db_kalinlik, boy, en, kanalli
                
                # ============================================================
                # ÖNCELİK: Malzeme kalınlığına göre genel tip belirle
                # Eğer malzeme 8mm ise → tip tespiti arkalık odaklı
                # Eğer malzeme 18mm ise → tip tespiti gövde odaklı
                # ============================================================
                
                # Malzeme 8mm (ince) ise - arkalık tipleri kontrol et
                if db_kalinlik <= ARKALIK_KALINLIK:
                    # Normal ARKALIK: (Yükseklik-18) x (Genişlik-18)
                    if modul_gen:
                        expected_arkalik_boy = yukseklik - 18
                        expected_arkalik_en = modul_gen - 18
                        if abs(boy - expected_arkalik_boy) <= TOLERANS and abs(en - expected_arkalik_en) <= TOLERANS:
                            return 'ARKALIK', db_kalinlik, boy, en, kanalli
                        
                        # ARKALIK (İÇERDE): (Yükseklik-37) x (Genişlik-37)
                        expected_arkalik_icerde_boy = yukseklik - 37
                        expected_arkalik_icerde_en = modul_gen - 37
                        if abs(boy - expected_arkalik_icerde_boy) <= TOLERANS and abs(en - expected_arkalik_icerde_en) <= TOLERANS:
                            return 'ARKALIK (İÇERDE)', db_kalinlik, boy, en, kanalli
                    
                    # Genel arkalık (modül genişliği bilinmiyorsa)
                    return 'ARKALIK', db_kalinlik, boy, en, kanalli
                
                # ============================================================
                # PARÇA TİPİ TESPİTİ - yukseklik ve derinlik değerlerini kullan
                # (cabinet_settings veya custom_module'den alındı)
                # ============================================================
                
                # YAN kontrolü - Yükseklik x Derinlik
                if abs(boy - yukseklik) <= TOLERANS and abs(en - derinlik) <= TOLERANS:
                    return 'YAN', db_kalinlik, boy, en, kanalli
                
                # ALT-ÜST: (Genişlik-36) x (Derinlik-1)
                if modul_gen:
                    expected_alt_ust_boy = modul_gen - 36
                    expected_alt_ust_en = derinlik - 1
                    
                    if abs(boy - expected_alt_ust_boy) <= TOLERANS and abs(en - expected_alt_ust_en) <= TOLERANS:
                        return 'ALT-ÜST', db_kalinlik, boy, en, kanalli
                
                # SABİT: (Genişlik-36) x (Derinlik-23)
                if modul_gen:
                    expected_sabit_boy = modul_gen - 36
                    expected_sabit_en = derinlik - 23
                    
                    if abs(boy - expected_sabit_boy) <= TOLERANS and abs(en - expected_sabit_en) <= TOLERANS:
                        return 'SABİT', db_kalinlik, boy, en, kanalli
                
                # RAF: (Genişlik-37) x (Derinlik-50 veya 40)
                # RAF ASLA KANALLI OLMAZ
                if modul_gen and not kanalli:
                    expected_raf_boy = modul_gen - 37
                    
                    # Dolap tipine göre raf düşümü: alt=50, üst=40
                    raf_dusum = 50 if dolap_tipi == 'alt' or dolap_tipi == 'boy' else 40
                    expected_raf_en = derinlik - raf_dusum
                    
                    if abs(boy - expected_raf_boy) <= TOLERANS and abs(en - expected_raf_en) <= TOLERANS:
                        if dolap_tipi == 'ust':
                            return 'RAF (ÜST)', db_kalinlik, boy, en, kanalli
                        else:
                            return 'RAF', db_kalinlik, boy, en, kanalli
                
                # ============================================================
                # MODÜL GENİŞLİĞİ YOKSA - SABİT ÖLÇÜLERLE KONTROL
                # ============================================================
                
                # ALT-ÜST kontrolü - EN değeri 579 veya 329 civarı
                if abs(en - 579) <= TOLERANS or abs(en - 329) <= TOLERANS:
                    return 'ALT-ÜST', db_kalinlik, boy, en, kanalli
                
                # SABİT kontrolü - EN değeri 557 veya 307 civarı
                if abs(en - 557) <= TOLERANS or abs(en - 307) <= TOLERANS:
                    return 'SABİT', db_kalinlik, boy, en, kanalli
                
                # RAF kontrolü - RAF ASLA KANALLI OLMAZ
                if not kanalli:
                    if abs(en - 530) <= TOLERANS:
                        return 'RAF', db_kalinlik, boy, en, kanalli
                    if abs(en - 290) <= TOLERANS:
                        return 'RAF (ÜST)', db_kalinlik, boy, en, kanalli
                
                # KAYIT/KUŞAK - Bir kenar 80-140mm civarı
                if 80 <= en <= 140 or 80 <= boy <= 140:
                    return 'KAYIT/KUŞAK', db_kalinlik, boy, en, kanalli
                
                # DİĞER - Hiçbirine uymayan
                return 'DİĞER', db_kalinlik, boy, en, kanalli

            # Parçaları işle - KANALLI bilgisini de ekle
            results = []
            for idx, row in df.iterrows():
                parca_tipi, kalinlik, boy, en, kanalli = determine_part_type_with_module(row)
                results.append({
                    'KALINLIK': kalinlik,
                    'MALZEME': row['MALZEME'],
                    'BOY': boy,
                    'EN': en,
                    'PARÇA TİPİ': parca_tipi,
                    'KANALLI': kanalli,  # Kanallı bilgisi eklendi
                    'ADET': row['ADET']
                })
            
            result_df = pd.DataFrame(results)
            
            # Gruplama - KANALLI bilgisini de dahil et
            group_cols = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ', 'KANALLI']
            summary = result_df.groupby(group_cols).agg({'ADET': 'sum'}).reset_index()
            
            # Ayır
            body_df = summary[summary['KALINLIK'] > ARKALIK_KALINLIK].sort_values(by=['MALZEME', 'KALINLIK', 'PARÇA TİPİ', 'BOY'])
            thin_df = summary[summary['KALINLIK'] <= ARKALIK_KALINLIK].sort_values(by=['MALZEME', 'PARÇA TİPİ', 'BOY'])

            return {
                'success': True,
                'job_no': job_no,
                'total_parts': int(summary['ADET'].sum()),
                'material_count': len(result_df['MALZEME'].unique()),
                'type_count': len(summary['PARÇA TİPİ'].unique()),
                'body': body_df.to_dict('records') if not body_df.empty else [],
                'thin': thin_df.to_dict('records') if not thin_df.empty else []
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def merge_jobs(self, jobs: List[Dict], output_path: str) -> Dict:
        if not PANDAS_AVAILABLE:
            return {'success': False, 'error': 'pandas kütüphanesi yüklü değil!'}

        try:
            all_body = []
            all_thin = []

            for job in jobs:
                results = job.get('results', {})
                all_body.extend(results.get('body', []))
                all_thin.extend(results.get('thin', []))

            if not all_body and not all_thin:
                return {'success': False, 'error': 'Birleştirilecek veri bulunamadı'}

            # Combine and group
            body_df = pd.DataFrame(all_body) if all_body else pd.DataFrame()
            thin_df = pd.DataFrame(all_thin) if all_thin else pd.DataFrame()

            if not body_df.empty:
                group_cols = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ']
                if 'MODÜL TİPİ' in body_df.columns:
                    group_cols.append('MODÜL TİPİ')
                if 'POZ' in body_df.columns:
                    group_cols.append('POZ')
                body_df = body_df.groupby(group_cols).agg({'ADET': 'sum'}).reset_index()
                body_df = body_df.sort_values(by=['MALZEME', 'PARÇA TİPİ', 'BOY'])

            if not thin_df.empty:
                group_cols = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ']
                if 'MODÜL TİPİ' in thin_df.columns:
                    group_cols.append('MODÜL TİPİ')
                if 'POZ' in thin_df.columns:
                    group_cols.append('POZ')
                thin_df = thin_df.groupby(group_cols).agg({'ADET': 'sum'}).reset_index()
                thin_df = thin_df.sort_values(by=['MALZEME', 'PARÇA TİPİ', 'BOY'])

            # Export with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                if not body_df.empty:
                    body_df.to_excel(writer, sheet_name='Gövde', index=False)
                    worksheet = writer.sheets['Gövde']
                    for idx, col in enumerate(body_df.columns):
                        max_len = max(body_df[col].astype(str).str.len().max(), len(col)) + 2
                        worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 30)
                        
                if not thin_df.empty:
                    thin_df.to_excel(writer, sheet_name='İnce', index=False)
                    worksheet = writer.sheets['İnce']
                    for idx, col in enumerate(thin_df.columns):
                        max_len = max(thin_df[col].astype(str).str.len().max(), len(col)) + 2
                        worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 30)

            total_parts = 0
            if not body_df.empty:
                total_parts += int(body_df['ADET'].sum())
            if not thin_df.empty:
                total_parts += int(thin_df['ADET'].sum())

            return {
                'success': True,
                'output_path': output_path,
                'total_parts': total_parts,
                'body': body_df.to_dict('records') if not body_df.empty else [],
                'thin': thin_df.to_dict('records') if not thin_df.empty else []
            }
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

# ============================================================
# API CLASS
# ============================================================

class Api:
    """JavaScript API interface"""

    def __init__(self):
        self.db = DatabaseManager()
        self.jsondata = JsonDataManager()
        self.usermgr = UserManager()
        self.current_file_paths = []
        self.custom_depths = {}
        self.custom_modules = {}  # POZ -> {genislik, yukseklik, derinlik}
        
        # Toplu dolap tipi ayarları (Alt/Üst/Boy dolap için varsayılan ölçüler)
        self.cabinet_settings = {
            'alt': {'yukseklik': 720, 'derinlik': 580},
            'ust': {'yukseklik': 720, 'derinlik': 330},
            'boy': {'yukseklik': 2100, 'derinlik': 580}
        }
        
        self.analyzer = ExcelAnalyzer(self.jsondata, self.custom_modules, self.cabinet_settings)

    # === User Management ===
    def login(self, username: str, password: str, remember_me: bool = False) -> Dict:
        result = self.usermgr.login(username, password, remember_me)
        
        # Beni Hatırla seçiliyse oturumu dosyaya kaydet
        if result.get('success') and remember_me and result.get('session_token'):
            self.usermgr.save_session_file(username, result['session_token'])
        
        return result
    
    def login_with_token(self, username: str, token: str) -> Dict:
        """Token ile otomatik giriş (Beni Hatırla)"""
        return self.usermgr.verify_session_token(username, token)
    
    def check_saved_session(self) -> Dict:
        """Dosyadan kayıtlı oturumu kontrol et ve doğrula"""
        return self.usermgr.load_session_file()
    
    def logout_user(self, username: str) -> Dict:
        """Çıkış yap ve token'ı temizle"""
        self.usermgr.clear_session_token(username)
        return {'success': True}

    def add_user(self, username: str, password: str, display_name: str = '', is_admin: bool = False) -> Dict:
        return self.usermgr.add_user(username, password, display_name, is_admin)

    def delete_user(self, username: str) -> Dict:
        return self.usermgr.delete_user(username)

    def get_users(self) -> List[Dict]:
        return self.usermgr.get_users_list()

    # === Settings Management ===
    def get_settings(self) -> Dict:
        return self.jsondata.get_settings()

    def save_settings(self, settings: Dict) -> Dict:
        if self.jsondata.save_settings(settings):
            return {'success': True}
        return {'success': False, 'error': 'Ayarlar kaydedilemedi'}

    def reset_settings(self) -> Dict:
        if self.jsondata.save_settings(self.jsondata.get_default_settings()):
            return {'success': True, 'settings': self.jsondata.get_default_settings()}
        return {'success': False, 'error': 'Ayarlar sıfırlanamadı'}

    # === Material Management ===
    def get_materials(self) -> Dict:
        return self.jsondata.get_materials()

    def save_material(self, code: str, thickness: int) -> Dict:
        if self.jsondata.save_material(code, thickness):
            return {'success': True}
        return {'success': False, 'error': 'Malzeme kaydedilemedi'}

    def delete_material(self, code: str) -> Dict:
        if self.jsondata.delete_material(code):
            return {'success': True}
        return {'success': False, 'error': 'Malzeme silinemedi'}

    def clear_materials(self) -> Dict:
        if self.jsondata.clear_materials():
            return {'success': True}
        return {'success': False, 'error': 'Malzemeler temizlenemedi'}

    def save_all_materials(self, materials: Dict) -> Dict:
        if self.jsondata.save_all_materials(materials):
            return {'success': True}
        return {'success': False, 'error': 'Malzemeler kaydedilemedi'}

    # === History Management ===
    def get_history(self) -> List[Dict]:
        return self.db.get_history()

    def delete_history(self, job_ids: List[int]) -> Dict:
        return self.db.delete_history(job_ids)

    def get_stats(self) -> Dict:
        return self.db.get_stats()

    # === File Operations ===
    def check_file(self, file_path: str) -> Dict:
        """Tek bir dosyayı kontrol et"""
        try:
            return self.analyzer.check_file(file_path)
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def check_all_files(self) -> dict:
        """Frontend'in toplu kontrol için çağırdığı API metodu."""
        try:
            if not self.current_file_paths:
                return {
                    'success': False,
                    'error': 'Kontrol edilecek dosya yok'
                }

            all_unknown = set()
            total_rows = 0

            for path in self.current_file_paths:
                check_result = self.analyzer.check_file(path)
                if not check_result.get('success'):
                    return check_result

                for code in check_result.get('unknown', []):
                    all_unknown.add(code)

                total_rows += check_result.get('row_count', 0)

            return {
                'success': True,
                'unknown': list(all_unknown),
                'total_rows': total_rows,
                'file_count': len(self.current_file_paths),
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}


    def handle_dropped_files(self, files_data: List[Dict]) -> Dict:
        """Handle files dropped onto the UI (content transfer)"""
        try:
            if not files_data:
                return {'success': False, 'error': 'Dosya verisi yok'}

            import base64
            import tempfile
            
            self.current_file_paths = []
            temp_dir = tempfile.gettempdir()

            for file_item in files_data:
                name = file_item.get('name', 'unknown.xlsx')
                content = file_item.get('content') # base64 string
                
                if not content:
                    continue
                
                # Basic validation
                if not name.lower().endswith(('.xlsx', '.xls', '.csv')):
                    continue

                # Remove header if present (e.g. "data:application/vnd.ms-excel;base64,")
                if ',' in content:
                    content = content.split(',')[1]
                
                # Save to temp file
                file_path = os.path.join(temp_dir, f"CORE_{name}")
                with open(file_path, 'wb') as f:
                    f.write(base64.b64decode(content))
                
                self.current_file_paths.append(file_path)

            if not self.current_file_paths:
                return {'success': False, 'error': 'Geçerli Excel dosyası işlenemedi'}

            files_info = []
            for path in self.current_file_paths:
                check_result = self.analyzer.check_file(path)
                if check_result['success']:
                    files_info.append({
                        'name': os.path.basename(path),
                        'path': path,
                        'job_no': check_result.get('job_no', '-'),
                        'unknown_materials': check_result.get('unknown', [])
                    })
            
            return {
                'success': True, 
                'files': files_info
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def select_file(self) -> Dict:
        """Select Excel file(s) for analysis"""
        try:
            result = webview.windows[0].create_file_dialog(
                webview.FileDialog.OPEN,
                allow_multiple=True,
                file_types=('Excel Files (*.xlsx;*.xls;*.csv)', 'All files (*.*)')
            )

            if not result:
                return {'success': False, 'error': 'Dosya seçilmedi'}

            # Handle single or multiple files
            self.current_file_paths = result if isinstance(result, (list, tuple)) else [result]

            files_info = []
            for path in self.current_file_paths:
                check_result = self.analyzer.check_file(path)
                if check_result['success']:
                    files_info.append({
                        'name': os.path.basename(path),
                        'path': path,
                        'job_no': check_result.get('job_no'),
                        'row_count': check_result.get('row_count'),
                        'material_count': check_result.get('material_count'),
                        'unknown_materials': check_result.get('unknown', [])
                    })

            return {
                'success': True,
                'files': files_info,
                'total_files': len(files_info),
                'has_unknown_materials': any(f['unknown_materials'] for f in files_info)
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def analyze_and_export(self, file_path: str) -> Dict:
        """Tek dosyayı analiz et ve Excel'e kaydet - tekrar işle için"""
        try:
            # Kayıt yeri seç
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            default_name = f"{base_name}_{timestamp}.xlsx"
            
            result_dialog = webview.windows[0].create_file_dialog(
                webview.FileDialog.SAVE,
                save_filename=default_name,
                file_types=('Excel Files (*.xlsx)', 'All files (*.*)')
            )
            
            if not result_dialog:
                return {'success': False, 'error': 'Kayıt yeri seçilmedi'}
            
            output_path = result_dialog if isinstance(result_dialog, str) else result_dialog[0]
            
            # Analiz et ve kaydet
            result = self.analyzer.analyze_and_export(file_path, output_path, self.custom_depths)
            
            if result['success']:
                # Geçmişe ekle
                job = {
                    'job_no': result.get('job_no') or f"JOB-{len(self.db.get_history()) + 1:04d}",
                    'date': datetime.now().strftime("%Y-%m-%d %H:%M"),
                    'file_name': os.path.basename(file_path),
                    'file_path': file_path,
                    'output_path': output_path,
                    'stats': {
                        'parts': result['total_parts'],
                        'materials': result['material_count'],
                        'types': result['type_count']
                    },
                    'results': {
                        'body': result['body'],
                        'thin': result['thin']
                    }
                }
                self.db.add_history(job)
            
            return result
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def analyze_file(self, file_index: int = 0) -> Dict:
        """Analyze a specific file - sadece analiz et, Excel'e kaydetme"""
        if not self.current_file_paths:
            return {'success': False, 'error': 'Dosya seçilmedi'}

        if file_index >= len(self.current_file_paths):
            return {'success': False, 'error': 'Geçersiz dosya indeksi'}

        try:
            path = self.current_file_paths[file_index]
            
            # Analiz et ama Excel'e kaydetme - sadece sonuçları döndür
            result = self.analyzer.analyze_only(path, self.custom_depths)
            
            return result
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def analyze_all_files(self) -> Dict:
        """Analyze all files separately"""
        if not self.current_file_paths:
            return {'success': False, 'error': 'Dosya seçilmedi'}

        try:
            # Get output directory
            result = webview.windows[0].create_file_dialog(webview.FileDialog.FOLDER)
            if not result:
                return {'success': False, 'error': 'Kayıt yeri seçilmedi'}

            output_dir = result if isinstance(result, str) else result[0]

            results = []
            errors = []

            for path in self.current_file_paths:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_name = os.path.splitext(os.path.basename(path))[0]
                output_path = os.path.join(output_dir, f"{base_name}_{timestamp}.xlsx")

                result = self.analyzer.analyze_and_export(path, output_path, self.custom_depths)

                if result['success']:
                    # Add to history
                    job = {
                        'job_no': result.get('job_no') or f"JOB-{len(self.db.get_history()) + 1:04d}",
                        'date': datetime.now().strftime("%Y-%m-%d %H:%M"),
                        'file_name': os.path.basename(path),
                        'file_path': path,
                        'output_path': output_path,
                        'stats': {
                            'parts': result['total_parts'],
                            'materials': result['material_count'],
                            'types': result['type_count']
                        },
                        'results': {
                            'body': result['body'],
                            'thin': result['thin']
                        }
                    }
                    self.db.add_history(job)
                    results.append({'file': os.path.basename(path), 'output': output_path, 'parts': result['total_parts']})
                else:
                    errors.append({'file': os.path.basename(path), 'error': result.get('error', 'Bilinmeyen hata')})

            return {
                'success': len(errors) == 0,
                'results': results,
                'errors': errors,
                'total_files': len(self.current_file_paths),
                'successful': len(results),
                'failed': len(errors)
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def merge_jobs(self, job_ids: List[int]) -> Dict:
        jobs = self.db.get_history_by_ids(job_ids)
        if not jobs:
            return {'success': False, 'error': 'Seçilen işler bulunamadı'}

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"BirlesikListe_{timestamp}.xlsx"

            result = webview.windows[0].create_file_dialog(
                webview.FileDialog.SAVE,
                save_filename=default_name,
                file_types=('Excel Files (*.xlsx)', 'All files (*.*)')
            )

            if not result:
                return {'success': False, 'error': 'Kayıt yeri seçilmedi'}

            output_path = result if isinstance(result, str) else result[0]

            return self.analyzer.merge_jobs(jobs, output_path)
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def open_file(self, file_path: str) -> Dict:
        """Open file with default application"""
        try:
            if not os.path.exists(file_path):
                return {'success': False, 'error': 'Dosya bulunamadı'}

            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', file_path])
            else:
                subprocess.run(['xdg-open', file_path])

            return {'success': True}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    # === Backup ===
    def create_backup(self) -> Dict:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"core_backup_{timestamp}.json"

            result = webview.windows[0].create_file_dialog(
                webview.FileDialog.SAVE,
                save_filename=default_name,
                file_types=('JSON Files (*.json)', 'All files (*.*)')
            )

            if not result:
                return {'success': False, 'error': 'Kayıt yeri seçilmedi'}

            output_path = result if isinstance(result, str) else result[0]

            data = {
                'version': Config.APP_VERSION,
                'created_at': datetime.now().isoformat(),
                'settings': self.jsondata.get_settings(),
                'materials': self.jsondata.get_materials(),
                'learned_parts': self.jsondata.get_learned_parts(),
                'history': self.db.get_history(1000)
            }

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return {'success': True, 'path': output_path}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def restore_backup(self) -> Dict:
        try:
            result = webview.windows[0].create_file_dialog(
                webview.OPEN_DIALOG,
                file_types=('JSON Files (*.json)', 'All files (*.*)')
            )

            if not result:
                return {'success': False, 'error': 'Dosya seçilmedi'}

            backup_path = result[0] if isinstance(result, tuple) else result

            with open(backup_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if 'settings' in data:
                self.jsondata.save_settings(data['settings'])
            if 'materials' in data:
                self.jsondata.save_all_materials(data['materials'])
            if 'learned_parts' in data:
                self.jsondata._write_json('learned_parts', data['learned_parts'])

            return {'success': True}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    # === Learned Parts (Öğrenme Sistemi) ===
    def save_learned_parts(self, parts: List[Dict]) -> Dict:
        """Öğrenilen parça tiplerini kaydet"""
        try:
            if self.jsondata.save_learned_parts_bulk(parts):
                return {'success': True, 'count': len(parts)}
            return {'success': False, 'error': 'Kurallar kaydedilemedi'}
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def get_learned_parts(self) -> Dict:
        """Öğrenilen parça tiplerini getir"""
        return self.jsondata.get_learned_parts()
    
    def clear_learned_parts(self) -> Dict:
        """Tüm öğrenilen kuralları sil"""
        if self.jsondata.clear_learned_parts():
            return {'success': True}
        return {'success': False, 'error': 'Kurallar silinemedi'}

    # === Module Depth (Modül Derinlik) ===
    def get_modules(self, file_path: str) -> Dict:
        """Excel dosyasından modül bilgilerini çıkar"""
        try:
            if not PANDAS_AVAILABLE:
                return {'success': False, 'error': 'pandas yüklü değil'}
            
            df = pd.read_excel(file_path)
            
            modules = {}
            
            # Info4 (POZ) ve Info5 (Modül adı) sütunlarını bul - hem §542 hem #8542
            poz_col = None
            modul_col = None
            
            for col in df.columns:
                col_str = str(col)
                if 'Info4' in col_str:
                    poz_col = col
                if 'Info5' in col_str:
                    modul_col = col
            
            print(f"get_modules - poz_col: {poz_col}, modul_col: {modul_col}")
            
            if not poz_col or not modul_col:
                return {'success': False, 'error': 'Dosyada modül bilgisi bulunamadı (Info4/Info5 sütunları yok)'}
            
            for idx, row in df.iterrows():
                poz = str(row.get(poz_col, '')).strip()
                modul_adi = str(row.get(modul_col, '')).strip()
                
                if not poz or poz == 'nan' or poz in modules:
                    continue
                
                # Modül tipini belirle
                modul_lower = modul_adi.lower()
                if 'üst' in modul_lower or 'ust' in modul_lower:
                    tip = 'ust'
                    varsayilan_derinlik = 330
                elif 'boy' in modul_lower:
                    tip = 'boy'
                    varsayilan_derinlik = 580
                else:
                    tip = 'alt'
                    varsayilan_derinlik = 580
                
                # Genişlik - modül adından çıkar
                cm_match = re.search(r'(\d+)\s*cm', modul_adi)
                genislik_cm = int(cm_match.group(1)) if cm_match else None
                
                modules[poz] = {
                    'ad': modul_adi,
                    'tip': tip,
                    'genislik_cm': genislik_cm,
                    'varsayilan_derinlik': varsayilan_derinlik,
                    'ozel_derinlik': self.custom_depths.get(poz)
                }
            
            print(f"Found {len(modules)} modules")
            return {'success': True, 'modules': modules}
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}
    
    def set_custom_depth(self, poz: str, depth: int) -> Dict:
        """POZ için özel derinlik ayarla"""
        self.custom_depths[poz] = depth
        return {'success': True}
    
    def set_custom_module(self, poz: str, values: Dict) -> Dict:
        """POZ için özel modül değerleri ayarla (genislik, yukseklik, derinlik)"""
        if not hasattr(self, 'custom_modules'):
            self.custom_modules = {}
        self.custom_modules[poz] = values
        # Eski uyumluluk için derinliği de ayarla
        if 'derinlik' in values:
            self.custom_depths[poz] = values['derinlik']
        return {'success': True}
    
    def get_custom_modules(self) -> Dict:
        """Özel modül değerlerini getir"""
        if not hasattr(self, 'custom_modules'):
            self.custom_modules = {}
        return self.custom_modules
    
    def clear_custom_depths(self) -> Dict:
        """Tüm özel ayarları temizle"""
        self.custom_depths = {}
        if hasattr(self, 'custom_modules'):
            self.custom_modules = {}
        return {'success': True}
    
    def get_custom_depths(self) -> Dict:
        """Özel derinlikleri getir"""
        return self.custom_depths
    
    # === Cabinet Settings (Toplu Dolap Ayarları) ===
    def get_cabinet_settings(self) -> Dict:
        """Toplu dolap tipi ayarlarını getir (alt/üst/boy)"""
        return self.cabinet_settings
    
    def set_cabinet_settings(self, cabinet_type: str, values: Dict) -> Dict:
        """
        Toplu dolap ayarlarını güncelle.
        cabinet_type: 'alt', 'ust', veya 'boy'
        values: {'yukseklik': int, 'derinlik': int}
        """
        if cabinet_type not in ['alt', 'ust', 'boy']:
            return {'success': False, 'error': f'Geçersiz dolap tipi: {cabinet_type}'}
        
        if 'yukseklik' in values:
            self.cabinet_settings[cabinet_type]['yukseklik'] = int(values['yukseklik'])
        if 'derinlik' in values:
            self.cabinet_settings[cabinet_type]['derinlik'] = int(values['derinlik'])
        
        # Analyzer'daki ayarları da güncelle
        self.analyzer.cabinet_settings = self.cabinet_settings
        
        return {'success': True, 'cabinet_settings': self.cabinet_settings}
    
    def reset_cabinet_settings(self) -> Dict:
        """Dolap ayarlarını varsayılana sıfırla"""
        self.cabinet_settings = {
            'alt': {'yukseklik': 720, 'derinlik': 580},
            'ust': {'yukseklik': 720, 'derinlik': 330},
            'boy': {'yukseklik': 2100, 'derinlik': 580}
        }
        self.analyzer.cabinet_settings = self.cabinet_settings
        return {'success': True, 'cabinet_settings': self.cabinet_settings}

    # === Export Edited Results ===
    def export_edited_results(self, body_data: List[Dict], thin_data: List[Dict], job_no: str = None) -> Dict:
        """Düzenlenmiş sonuçları Excel'e kaydet"""
        try:
            if not PANDAS_AVAILABLE:
                return {'success': False, 'error': 'pandas kütüphanesi yüklü değil!'}
            
            # Kayıt yeri seç
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"KesimListesi_{job_no or timestamp}.xlsx"

            result = webview.windows[0].create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename=default_name,
                file_types=('Excel Files (*.xlsx)', 'All files (*.*)')
            )

            if not result:
                return {'success': False, 'error': 'Kayıt yeri seçilmedi'}

            output_path = result if isinstance(result, str) else result[0]
            
            # DataFrame'lere dönüştür
            all_data_df = pd.DataFrame(body_data + thin_data) if (body_data or thin_data) else pd.DataFrame()
            
            # Kanallı parçaların tipine (K) ekle
            def add_kanalli_to_type(df):
                if df.empty:
                    return df
                if 'KANALLI' in df.columns and 'PARÇA TİPİ' in df.columns:
                    df['PARÇA TİPİ'] = df.apply(
                        lambda row: f"{row['PARÇA TİPİ']} (K)" if row.get('KANALLI', False) else row['PARÇA TİPİ'],
                        axis=1
                    )
                return df
            
            all_data_df = add_kanalli_to_type(all_data_df)
            
            # Kalınlığa göre 3 tabloya ayır
            if not all_data_df.empty and 'KALINLIK' in all_data_df.columns:
                df_18mm = all_data_df[all_data_df['KALINLIK'] == 18].copy()
                df_16mm = all_data_df[all_data_df['KALINLIK'] == 16].copy()
                df_8mm = all_data_df[all_data_df['KALINLIK'] <= 8].copy()
            else:
                df_18mm = pd.DataFrame()
                df_16mm = pd.DataFrame()
                df_8mm = pd.DataFrame()
            
            # PARÇA TİPİ'ne göre sırala - aynı tipler alt alta gelsin
            part_type_order = [
                'YAN', 'YAN (K)',
                'ALT-ÜST', 'ALT-ÜST (K)',
                'SABİT', 'SABİT (K)',
                'RAF', 'RAF (K)',
                'RAF (ÜST)', 'RAF (ÜST) (K)',
                'KAYIT/KUŞAK', 'KAYIT/KUŞAK (K)',
                'ÇEKMECE YANI', 'ÇEKMECE YANI (K)',
                'ARKALIK', 'ARKALIK (K)',
                'ARKALIK (İÇERDE)', 'ARKALIK (İÇERDE) (K)',
                'DİĞER', 'DİĞER (K)'
            ]
            
            def sort_by_part_type(df):
                if df.empty or 'PARÇA TİPİ' not in df.columns:
                    return df
                df['_sort_order'] = df['PARÇA TİPİ'].apply(
                    lambda x: part_type_order.index(x) if x in part_type_order else 999
                )
                df = df.sort_values(by=['_sort_order', 'MALZEME', 'BOY', 'EN'])
                df = df.drop(columns=['_sort_order'])
                return df
            
            df_18mm = sort_by_part_type(df_18mm)
            df_16mm = sort_by_part_type(df_16mm)
            df_8mm = sort_by_part_type(df_8mm)
            
            # Excel'e yaz - TEK SHEET'TE 3 TABLO YAN YANA
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Kesim Listesi'
            
            # Stil tanımları - 3 farklı renk
            header_font = Font(bold=True, color='FFFFFF')
            header_fill_18mm = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Mavi
            header_fill_16mm = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')  # Mor
            header_fill_8mm = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')   # Yeşil
            title_font = Font(bold=True, size=14)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # KANALLI sütununu çıkar (artık parça tipine eklendi)
            for df in [df_18mm, df_16mm, df_8mm]:
                if not df.empty and 'KANALLI' in df.columns:
                    df.drop(columns=['KANALLI'], inplace=True)
            
            # Sütun sıralaması
            column_order = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ', 'ADET']
            cols_count = 6
            table_gap = 1  # Tablolar arası boşluk
            
            def write_table(ws, start_col, df, title, header_fill):
                """Tek bir tabloyu belirtilen sütundan başlayarak yaz"""
                current_row = 1
                
                # Zebra efekti için gri renk
                gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                # Başlık - ORTALI
                ws.cell(row=current_row, column=start_col, value=title)
                ws.cell(row=current_row, column=start_col).font = title_font
                ws.cell(row=current_row, column=start_col).alignment = Alignment(horizontal='center')
                if cols_count > 1:
                    ws.merge_cells(start_row=current_row, start_column=start_col, 
                                  end_row=current_row, end_column=start_col + cols_count - 1)
                
                current_row += 1
                
                # Header satırı
                header_names = ['KALINLIK', 'MALZEME', 'BOY', 'EN', 'PARÇA TİPİ', 'ADET']
                for col_idx, col_name in enumerate(header_names):
                    cell = ws.cell(row=current_row, column=start_col + col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                current_row += 1
                
                # Veri satırları - ZEBRA EFEKTİ (beyaz/gri)
                if not df.empty:
                    df_ordered = df.copy()
                    existing_cols = [c for c in column_order if c in df_ordered.columns]
                    df_ordered = df_ordered[existing_cols]
                    
                    row_index = 0
                    for row_data in df_ordered.values:
                        for col_idx, value in enumerate(row_data):
                            cell = ws.cell(row=current_row, column=start_col + col_idx, value=value)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                            # Çift satırlar gri (0, 2, 4...), tek satırlar beyaz
                            if row_index % 2 == 1:
                                cell.fill = gray_fill
                        current_row += 1
                        row_index += 1
                else:
                    ws.cell(row=current_row, column=start_col, value='(Veri yok)')
                    current_row += 1
                
                return current_row
            
            # ========== 1. TABLO: 18mm PARÇALAR (Mavi) ==========
            col_18mm_start = 1
            write_table(worksheet, col_18mm_start, df_18mm, '18mm Parçalar', header_fill_18mm)
            
            # ========== 2. TABLO: 16mm PARÇALAR (Mor) ==========
            col_16mm_start = col_18mm_start + cols_count + table_gap
            write_table(worksheet, col_16mm_start, df_16mm, '16mm Parçalar', header_fill_16mm)
            
            # ========== 3. TABLO: 8mm PARÇALAR (Yeşil) ==========
            col_8mm_start = col_16mm_start + cols_count + table_gap
            write_table(worksheet, col_8mm_start, df_8mm, '8mm Parçalar', header_fill_8mm)
            
            # Sütun genişliklerini ayarla
            # Sütun sırası: KALINLIK, MALZEME, BOY, EN, PARÇA TİPİ, ADET
            # 18mm ve 16mm için genişlikler
            widths_normal = [9, None, 9.5, 9.5, 14.2, 9]  # None = otomatik
            # 8mm için genişlikler (PARÇA TİPİ farklı)
            widths_8mm = [9, None, 9.5, 9.5, 15.7, 9]
            
            def set_column_widths(start_col, df, widths):
                for i, width in enumerate(widths):
                    col_letter = get_column_letter(start_col + i)
                    if width is None:
                        # Otomatik: içeriğe göre ayarla
                        max_len = len('MALZEME')  # Header uzunluğu
                        if not df.empty and 'MALZEME' in df.columns:
                            max_content = df['MALZEME'].astype(str).str.len().max()
                            max_len = max(max_len, max_content)
                        worksheet.column_dimensions[col_letter].width = max_len + 2
                    else:
                        worksheet.column_dimensions[col_letter].width = width
            
            set_column_widths(col_18mm_start, df_18mm, widths_normal)
            set_column_widths(col_16mm_start, df_16mm, widths_normal)
            set_column_widths(col_8mm_start, df_8mm, widths_8mm)
            
            # Toplam parça sayısı hesapla
            total_parts = 0
            if not df_18mm.empty and 'ADET' in df_18mm.columns:
                total_parts += int(df_18mm['ADET'].sum())
            if not df_16mm.empty and 'ADET' in df_16mm.columns:
                total_parts += int(df_16mm['ADET'].sum())
            if not df_8mm.empty and 'ADET' in df_8mm.columns:
                total_parts += int(df_8mm['ADET'].sum())
            
            # Excel dosyasını kaydet
            workbook.save(output_path)
            
            # History'ye ekle
            if self.current_file_paths:
                file_path = self.current_file_paths[0]
                file_name = os.path.basename(file_path)
            else:
                file_path = ''
                file_name = 'Manuel Düzenleme'
            
            # Tüm malzemeleri topla
            all_materials = set()
            all_types = set()
            for df in [df_18mm, df_16mm, df_8mm]:
                if not df.empty:
                    if 'MALZEME' in df.columns:
                        all_materials.update(df['MALZEME'].tolist())
                    if 'PARÇA TİPİ' in df.columns:
                        all_types.update(df['PARÇA TİPİ'].tolist())
            
            job = {
                'job_no': job_no or f"JOB-{len(self.db.get_history()) + 1:04d}",
                'date': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'file_name': file_name,
                'file_path': file_path,
                'output_path': output_path,
                'stats': {
                    'parts': total_parts,
                    'materials': len(all_materials),
                    'types': len(all_types)
                },
                'results': {
                    'body': body_data,
                    'thin': thin_data
                }
            }
            self.db.add_history(job)
            
            return {
                'success': True,
                'output_path': output_path,
                'total_parts': total_parts
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def analyze_and_export_filtered(self, file_index: int, selected_types: List[str]) -> Dict:
        """
        Sadece seçilen parça tiplerini içeren filtrelenmiş analiz ve export.
        
        Args:
            file_index: Dosya indeksi
            selected_types: Seçilen parça tipleri listesi (örn: ['YAN', 'RAF', 'ALT-ÜST'])
        
        Returns:
            Dict: Başarı durumu ve çıktı yolu
        """
        try:
            if not self.current_file_paths:
                return {'success': False, 'error': 'Dosya seçilmedi'}
            
            if file_index >= len(self.current_file_paths):
                return {'success': False, 'error': 'Geçersiz dosya indeksi'}
            
            path = self.current_file_paths[file_index]
            
            # Önce normal analiz yap
            result = self.analyzer.analyze_only(path, self.custom_depths)
            
            if not result.get('success'):
                return result
            
            # Body ve thin verilerini filtrele
            body_data = result.get('body', [])
            thin_data = result.get('thin', [])
            
            # Sadece seçilen tipleri tut
            filtered_body = [row for row in body_data if row.get('PARÇA TİPİ') in selected_types]
            filtered_thin = [row for row in thin_data if row.get('PARÇA TİPİ') in selected_types]
            
            # Filtrelenmiş sonuçlarla export et
            job_no = result.get('job_no', '')
            export_result = self.export_edited_results(filtered_body, filtered_thin, job_no)
            
            if export_result.get('success'):
                export_result['filtered'] = True
                export_result['selected_types'] = selected_types
                export_result['total_parts'] = sum(row.get('ADET', 0) for row in filtered_body + filtered_thin)
            
            return export_result
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

# ============================================================
# MAIN APPLICATION
# ============================================================

def main():
    """Main entry point"""
    print(f"Starting {Config.APP_TITLE}...")
    print(f"Base path: {Config.get_base_path()}")
    print(f"App dir: {Config.get_app_dir()}")
    print(f"Data dir: {Config.get_data_dir()}")

    # Check for index.html
    html_path = Config.get_html_path()
    print(f"HTML path: {html_path}")

    if not html_path.exists():
        print(f"ERROR: index.html not found at {html_path}")
        # Try alternative paths
        alt_paths = [
            get_app_dir() / "index.html",
            Path(__file__).parent / "index.html",
            Path.cwd() / "index.html"
        ]
        for alt in alt_paths:
            print(f"Trying: {alt}")
            if alt.exists():
                html_path = alt
                print(f"Found at: {html_path}")
                break
        else:
            print("ERROR: Could not find index.html anywhere!")
            sys.exit(1)

    # Create API instance
    api = Api()

    # Create window
    window = webview.create_window(
        title=Config.APP_TITLE,
        url=str(html_path),
        js_api=api,
        width=Config.WINDOW_WIDTH,
        height=Config.WINDOW_HEIGHT,
        min_size=(Config.WINDOW_MIN_WIDTH, Config.WINDOW_MIN_HEIGHT),
        resizable=True,
        text_select=False,
        confirm_close=True
    )

    print("Window created, starting webview...")

    # Start webview (debug=False for production)
    webview.start(debug=False)

if __name__ == "__main__":
    main()